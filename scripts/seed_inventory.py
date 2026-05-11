"""
Seed diamond_inventory, metal_rates, pricing_settings from Price Chart_Karigar.xlsx.

Usage:
    python scripts/seed_inventory.py --excel "C:/path/to/Price Chart_Karigar.xlsx"
    python scripts/seed_inventory.py --excel "C:/path/to/Price Chart_Karigar.xlsx" --db-url postgresql://postgres:pass@localhost:5432/karigar_app
"""
import argparse
import asyncio
import logging
import os
import re
from datetime import datetime

import asyncpg
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger(__name__)


def _recover_date_encoded(val) -> str | None:
    """
    Excel auto-converts sieve/batch values like "3/8", "1/2", "2-6.5"
    into date objects.  Recover the original fraction/range text.
    e.g. datetime(2025, 3, 8) → "3/8"
         datetime(2025, 1, 2) → "1/2"
    """
    if isinstance(val, datetime):
        return f"{val.month}/{val.day}"
    return None


def safe_str(val) -> str | None:
    if val is None:
        return None
    recovered = _recover_date_encoded(val)
    if recovered:
        return recovered
    s = str(val).strip()
    return s if s else None


def safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def safe_float(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _normalize_sieve(val: str | None) -> str | None:
    """Strip leading + or - that the LGD sheet prepends to sieve size values."""
    if val and val[0] in ('+', '-'):
        return val[1:]
    return val


_LGD_TYPES = {"LGD_HPHT", "LGD_CVD"}
_NATURAL_TYPES = {"Natural Diamond", "Natural Fency Diamond"}


def _karigar_fallback(item_type: str, vendor_price: int | None) -> int:
    """
    Compute karigar price from vendor when the Excel formula cell has no cached
    value (file not saved by Excel after last edit).
    LGD markup: vendor × 1.10, rounded to nearest ₹100.
    Natural markup: vendor × 1.04, rounded to nearest ₹100.
    """
    if not vendor_price:
        return 0
    if item_type in _LGD_TYPES:
        return round(vendor_price * 1.10 / 100) * 100
    if item_type in _NATURAL_TYPES:
        return round(vendor_price * 1.04 / 100) * 100
    return 0


# ── Main sheets to ingest ─────────────────────────────────────────────────────

def parse_master_sheet(ws) -> list[tuple]:
    """Parse Master / LGD / RBC / Natural Fency sheet (cols A-M, row 2 onwards)."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_type = safe_str(row[0])
        if not item_type:
            continue
        vendor_price = safe_int(row[10])
        karigar_price = safe_int(row[12]) or _karigar_fallback(item_type, vendor_price)
        rows.append((
            item_type,
            safe_str(row[1]),                    # batch
            safe_str(row[2]),                    # colour
            safe_str(row[3]),                    # shape
            safe_str(row[4]),                    # mm_size
            _normalize_sieve(safe_str(row[5])),  # sieve_size — strip leading +/-
            safe_float(row[6]),                  # avg_weight_carat
            safe_str(row[7]),                    # karigar_color
            safe_str(row[8]),                    # karigar_purity
            safe_str(row[9]),                    # company_grade
            vendor_price,                        # vendor_price
            safe_str(row[11]),                   # vendor_name
            karigar_price,
        ))
    return rows


def parse_lgd_sheet(ws) -> list[tuple]:
    return parse_master_sheet(ws)


def parse_rbc_sheet(ws) -> list[tuple]:
    return parse_master_sheet(ws)


def parse_fancy_sheet(ws) -> list[tuple]:
    return parse_master_sheet(ws)


# ── DB helpers ────────────────────────────────────────────────────────────────

INSERT_INVENTORY = """
    INSERT INTO diamond_inventory
        (item_type, batch, colour, shape, mm_size, sieve_size,
         avg_weight_carat, karigar_color, karigar_purity, company_grade,
         vendor_price, vendor_name, karigar_price)
    VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13)
"""


async def seed_inventory(conn: asyncpg.Connection, wb) -> int:
    log.info("Truncating diamond_inventory …")
    await conn.execute("TRUNCATE diamond_inventory RESTART IDENTITY")

    all_rows: list[tuple] = []
    for sheet_name, parser in [
        ("Master",        parse_master_sheet),
        ("LGD",           parse_lgd_sheet),
        ("RBC",           parse_rbc_sheet),
        ("Natural Fency", parse_fancy_sheet),
    ]:
        if sheet_name not in wb.sheetnames:
            log.warning("Sheet '%s' not found — skipping", sheet_name)
            continue
        rows = parser(wb[sheet_name])
        log.info("  %-16s  %d rows", sheet_name, len(rows))
        all_rows.extend(rows)

    await conn.executemany(INSERT_INVENTORY, all_rows)
    log.info("Inserted %d total inventory rows", len(all_rows))
    return len(all_rows)


async def seed_metal_rates(conn: asyncpg.Connection, wb) -> None:
    if "Metal" not in wb.sheetnames:
        log.warning("Metal sheet not found — skipping metal rates")
        return

    ws = wb["Metal"]

    # First pass: read raw values. price_per_gram may be None when formula cells
    # have no cached value (file not opened/saved by Excel after last edit).
    raw = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        metal_type = safe_str(row[1])
        if not metal_type:
            continue
        raw.append({
            "metal_type":    metal_type.strip(),
            "purity":        safe_int(row[2]) or 0,
            "colour":        safe_str(row[3]) or "Yellow",
            "weight_gms":    safe_float(row[4]) or 10.0,   # col E: Weight/GMS reference
            "market_rate":   safe_float(row[5]),            # col F: rate per weight_gms grams
            "price_per_gram": safe_float(row[7]),           # col H: formula result
        })

    # Resolve 24kt gold base price; all other purities are derived from it.
    gold_24_ppg: float | None = None
    for r in raw:
        if r["metal_type"] == "Gold" and r["purity"] == 24:
            if r["price_per_gram"] is None and r["market_rate"] is not None:
                r["price_per_gram"] = r["market_rate"] / r["weight_gms"]
            gold_24_ppg = r["price_per_gram"]
            break

    records = []
    for r in raw:
        ppg = r["price_per_gram"]
        if ppg is None:
            if r["metal_type"] == "Gold" and gold_24_ppg is not None:
                # Excel formula: =H2 * purity_kt / 24
                ppg = gold_24_ppg * r["purity"] / 24
            elif r["metal_type"] == "Silver" and r["market_rate"] is not None:
                ppg = r["market_rate"] / r["weight_gms"]
        if ppg is None:
            log.warning("Cannot determine price_per_gram for %s %skt — skipping", r["metal_type"], r["purity"])
            continue
        records.append((r["metal_type"], r["purity"], r["colour"], r["market_rate"], ppg))

    await conn.executemany("""
        INSERT INTO metal_rates (metal_type, purity_kt, colour, market_rate_per_kg, price_per_gram)
        VALUES ($1, $2, $3, $4, $5)
        ON CONFLICT (metal_type, purity_kt, colour) DO UPDATE
            SET market_rate_per_kg = EXCLUDED.market_rate_per_kg,
                price_per_gram     = EXCLUDED.price_per_gram,
                updated_at         = NOW()
    """, records)
    log.info("Upserted %d metal rate rows", len(records))


async def seed_pricing_settings(conn: asyncpg.Connection, wb) -> None:
    if "Settings" not in wb.sheetnames:
        log.warning("Settings sheet not found — using defaults")
        return

    ws = wb["Settings"]
    mapping = {
        "Labour Per Gram": "labour_per_gram",
        "Markup %":        "markup_pct",
        "GST %":           "gst_pct",
    }
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        key_raw = safe_str(row[0])
        val = safe_float(row[1])
        if not key_raw or val is None:
            continue
        db_key = mapping.get(key_raw, re.sub(r"\s+", "_", key_raw).lower())
        await conn.execute("""
            INSERT INTO pricing_settings (key, value)
            VALUES ($1, $2)
            ON CONFLICT (key) DO UPDATE
                SET value = EXCLUDED.value, updated_at = NOW()
        """, db_key, val)
    log.info("Pricing settings seeded")


async def run(db_url: str, excel_path: str) -> None:
    # asyncpg uses plain postgres:// URLs
    url = db_url.replace("postgresql+asyncpg://", "postgresql://")
    log.info("Connecting to DB …")
    conn = await asyncpg.connect(url)
    try:
        log.info("Loading workbook: %s", excel_path)
        wb = load_workbook(excel_path, data_only=True)

        await seed_inventory(conn, wb)
        await seed_metal_rates(conn, wb)
        await seed_pricing_settings(conn, wb)
        log.info("Seeding complete.")
    finally:
        await conn.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Seed Karigar inventory from Excel")
    parser.add_argument("--excel", required=True, help="Path to Price Chart_Karigar.xlsx")
    parser.add_argument(
        "--db-url",
        default=os.getenv(
            "DATABASE_URL",
            "postgresql://postgres:postgres@localhost:5432/karigar_app",
        ),
    )
    args = parser.parse_args()
    asyncio.run(run(args.db_url, args.excel))

from __future__ import annotations

import logging
import os
import re
import threading
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_WORKBOOK_PATH = os.getenv(
    "KARIGAR_PRICING_WORKBOOK_PATH",
    r"C:\Users\Ayush Tripathi\Downloads\Price Chart_Karigar.xlsx",
)
logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class PricingSettings:
    labour_per_gram: Decimal
    markup: Decimal
    gst: Decimal


@dataclass(frozen=True)
class MasterPriceRow:
    source_row: int
    item_name: str
    batch: str
    colour: str
    shape: str
    mm_size: str
    sieve_size: str
    carat_weight: Decimal | None
    karigar_colour: str
    karigar_purity: str
    company_grade: str
    vendor_price: Decimal
    vendor_name: str
    karigar_price: Decimal


class PriceNotFoundError(Exception):
    pass


class WorkbookNotConfiguredError(Exception):
    pass


_SLASH_PAIR_RE = re.compile(r"^\s*(\d+)\s*/\s*(\d+)\s*$")


class KarigarPricingEngine:
    def __init__(self, workbook_path: str = DEFAULT_WORKBOOK_PATH) -> None:
        self.workbook_path = Path(workbook_path)
        self._is_loaded = False
        self._settings = PricingSettings(
            labour_per_gram=Decimal("1000"),
            markup=Decimal("0.15"),
            gst=Decimal("0.03"),
        )
        self._master_rows: list[MasterPriceRow] = []
        self._lookup: dict[tuple[str, str, str, str, str, str], MasterPriceRow] = {}
        self._duplicate_key_conflicts = 0
        self._load_lock = threading.Lock()
        self._dropdown_options: dict[str, tuple[str, ...]] = {
            "item_names": (),
            "batches": (),
            "colours": (),
            "clarities": (),
            "shapes": (),
            "sieve_sizes": (),
        }

    @property
    def settings(self) -> PricingSettings:
        self._ensure_loaded()
        return self._settings

    @property
    def master_row_count(self) -> int:
        self._ensure_loaded()
        return len(self._master_rows)

    @property
    def duplicate_key_conflicts(self) -> int:
        self._ensure_loaded()
        return self._duplicate_key_conflicts

    def _ensure_loaded(self) -> None:
        with self._load_lock:
            if self._is_loaded:
                return
            if not self.workbook_path.exists():
                raise WorkbookNotConfiguredError(
                    f"Workbook not found at '{self.workbook_path}'. Set KARIGAR_PRICING_WORKBOOK_PATH."
                )

            workbook_values = load_workbook(filename=str(self.workbook_path), data_only=True)
            workbook_raw = load_workbook(filename=str(self.workbook_path), data_only=False)
            self._settings = self._read_settings(workbook_values)
            self._master_rows = self._read_master_rows(workbook_values, workbook_raw)
            self._lookup = {}
            conflict_keys: set[tuple[str, str, str, str, str, str]] = set()
            for row in self._master_rows:
                key = (
                    row.item_name,
                    row.batch,
                    row.karigar_colour,
                    row.karigar_purity,
                    row.shape,
                    row.sieve_size,
                )
                existing = self._lookup.get(key)
                if existing is not None and row.karigar_price != existing.karigar_price:
                    conflict_keys.add(key)
                    logger.warning(
                        "Conflicting karigar_price for lookup key %s between row %s (price=%s) "
                        "and row %s (price=%s). Keeping the higher karigar_price.",
                        key,
                        existing.source_row,
                        existing.karigar_price,
                        row.source_row,
                        row.karigar_price,
                    )
                if existing is None or row.karigar_price > existing.karigar_price:
                    self._lookup[key] = row
            self._duplicate_key_conflicts = len(conflict_keys)
            self._dropdown_options = {
                "item_names": tuple(sorted({row.item_name for row in self._master_rows if row.item_name})),
                "batches": tuple(sorted({row.batch for row in self._master_rows if row.batch})),
                "colours": tuple(sorted({row.karigar_colour for row in self._master_rows if row.karigar_colour})),
                "clarities": tuple(sorted({row.karigar_purity for row in self._master_rows if row.karigar_purity})),
                "shapes": tuple(sorted({row.shape for row in self._master_rows if row.shape})),
                "sieve_sizes": tuple(sorted({row.sieve_size for row in self._master_rows if row.sieve_size})),
            }
            workbook_values.close()
            workbook_raw.close()
            self._is_loaded = True

    def reload(self) -> None:
        with self._load_lock:
            self._is_loaded = False
            self._master_rows = []
            self._lookup = {}
            self._duplicate_key_conflicts = 0
            self._dropdown_options = {
                "item_names": (),
                "batches": (),
                "colours": (),
                "clarities": (),
                "shapes": (),
                "sieve_sizes": (),
            }
        self._ensure_loaded()

    def get_karigar_price(
        self,
        item_name: str,
        batch: str,
        colour: str,
        clarity: str,
        shape: str,
        sieve_size: str,
    ) -> MasterPriceRow:
        self._ensure_loaded()
        key = self._build_lookup_key(
            item_name=item_name,
            batch=batch,
            colour=colour,
            clarity=clarity,
            shape=shape,
            sieve_size=sieve_size,
        )
        row = self._lookup.get(key)
        if row is None:
            raise PriceNotFoundError("Price Not Found")
        if row.karigar_price == Decimal("0"):
            raise PriceNotFoundError("Price Not Found")
        return row

    def get_lookup_debug(
        self,
        item_name: str,
        batch: str,
        colour: str,
        clarity: str,
        shape: str,
        sieve_size: str,
    ) -> dict[str, object]:
        self._ensure_loaded()
        key = self._build_lookup_key(
            item_name=item_name,
            batch=batch,
            colour=colour,
            clarity=clarity,
            shape=shape,
            sieve_size=sieve_size,
        )
        key_item_name = key[0]
        sample_rows = [
            row
            for row in self._master_rows
            if row.item_name == key_item_name
        ][:5]
        return {
            "normalized_key": {
                "item_name": key[0],
                "batch": key[1],
                "colour": key[2],
                "clarity": key[3],
                "shape": key[4],
                "sieve_size": key[5],
            },
            "lookup_index_size": len(self._lookup),
            "item_name_sample_rows": [
                {
                    "item_name": row.item_name,
                    "batch": row.batch,
                    "karigar_colour": row.karigar_colour,
                    "karigar_purity": row.karigar_purity,
                    "shape": row.shape,
                    "sieve_size": row.sieve_size,
                    "karigar_price": row.karigar_price,
                    "vendor_price": row.vendor_price,
                    "vendor_name": row.vendor_name,
                }
                for row in sample_rows
            ],
        }

    def get_quotation(
        self,
        gold_weight: Decimal,
        gold_rate: Decimal,
        diamond_value: Decimal | None = None,
        karigar_price_per_carat: Decimal | None = None,
        total_carat_weight: Decimal | None = None,
        labour_per_gram: Decimal | None = None,
        markup: Decimal | None = None,
        gst: Decimal | None = None,
    ) -> dict[str, Decimal]:
        cfg = self.settings
        labour_rate = cfg.labour_per_gram if labour_per_gram is None else labour_per_gram
        markup_rate = cfg.markup if markup is None else markup
        gst_rate = cfg.gst if gst is None else gst
        resolved_diamond_value = diamond_value
        if resolved_diamond_value is None and karigar_price_per_carat is not None and total_carat_weight is not None:
            resolved_diamond_value = karigar_price_per_carat * total_carat_weight
        if resolved_diamond_value is None:
            resolved_diamond_value = Decimal("0")

        gold_value = gold_weight * gold_rate
        labour = gold_weight * labour_rate
        subtotal = gold_value + labour + resolved_diamond_value
        after_markup = subtotal * (Decimal("1") + markup_rate)
        final_price = after_markup * (Decimal("1") + gst_rate)

        return {
            "gold_value": gold_value,
            "labour": labour,
            "diamond_value": resolved_diamond_value,
            "subtotal": subtotal,
            "after_markup": after_markup,
            "final_price": final_price,
        }

    def get_dropdown_options(self) -> dict[str, list[str]]:
        self._ensure_loaded()
        return {
            "item_names": list(self._dropdown_options["item_names"]),
            "batches": list(self._dropdown_options["batches"]),
            "colours": list(self._dropdown_options["colours"]),
            "clarities": list(self._dropdown_options["clarities"]),
            "shapes": list(self._dropdown_options["shapes"]),
            "sieve_sizes": list(self._dropdown_options["sieve_sizes"]),
        }

    def _build_lookup_key(
        self,
        item_name: str,
        batch: str,
        colour: str,
        clarity: str,
        shape: str,
        sieve_size: str,
    ) -> tuple[str, str, str, str, str, str]:
        return (
            self._stringify(item_name),
            self._normalize_batch_or_sieve(self._stringify(batch)),
            self._stringify(colour),
            self._stringify(clarity),
            self._stringify(shape),
            self._normalize_batch_or_sieve(self._stringify(sieve_size)),
        )

    def _read_settings(self, workbook) -> PricingSettings:
        ws = workbook["Settings"]
        return PricingSettings(
            labour_per_gram=self._to_decimal(ws["B1"].value, Decimal("1000")),
            markup=self._to_decimal(ws["B2"].value, Decimal("0.15")),
            gst=self._to_decimal(ws["B3"].value, Decimal("0.03")),
        )

    def _read_master_rows(self, workbook_values, workbook_raw) -> list[MasterPriceRow]:
        ws_values = workbook_values["Master"]
        ws_raw = workbook_raw["Master"]
        rows: list[MasterPriceRow] = []

        for idx in range(2, ws_values.max_row + 1):
            raw_item_name = ws_raw.cell(row=idx, column=1).value
            item_name = self._stringify(raw_item_name)
            if not item_name:
                continue

            rows.append(
                MasterPriceRow(
                    source_row=idx,
                    item_name=item_name,
                    # Keep text dimensions from the raw workbook so Excel-cast
                    # values like "1/4" survive as lookup-able text keys.
                    batch=self._normalize_batch_or_sieve(
                        self._stringify(ws_raw.cell(row=idx, column=2).value)
                    ),
                    colour=self._stringify(ws_values.cell(row=idx, column=3).value),
                    shape=self._stringify(ws_raw.cell(row=idx, column=4).value),
                    mm_size=self._stringify(ws_values.cell(row=idx, column=5).value),
                    sieve_size=self._normalize_batch_or_sieve(
                        self._stringify(ws_raw.cell(row=idx, column=6).value)
                    ),
                    # Numeric measures/prices must come from data_only workbook.
                    carat_weight=self._optional_decimal(ws_values.cell(row=idx, column=7).value),
                    karigar_colour=self._stringify(ws_raw.cell(row=idx, column=8).value),
                    karigar_purity=self._stringify(ws_raw.cell(row=idx, column=9).value),
                    company_grade=self._stringify(ws_raw.cell(row=idx, column=10).value),
                    vendor_price=self._to_decimal(ws_values.cell(row=idx, column=11).value, Decimal("0")),
                    vendor_name=self._stringify(ws_raw.cell(row=idx, column=12).value),
                    karigar_price=self._to_decimal(ws_values.cell(row=idx, column=13).value, Decimal("0")),
                )
            )

        return rows

    @staticmethod
    def _stringify(value: object) -> str:
        # Batch/sieve values like "2-6.5" and "1/4" must remain string-compared.
        if value is None:
            return ""
        if isinstance(value, datetime):
            return f"{value.month}/{value.day}"
        if isinstance(value, date):
            return f"{value.month}/{value.day}"
        return str(value).strip()

    @staticmethod
    def _normalize_batch_or_sieve(value: str) -> str:
        # Excel may parse "3/8" as August 3rd and re-serialize as "8/3"
        # (day/month). Sorting both numbers low/high makes both forms map to
        # the same canonical key. This is valid for this dataset because all
        # fraction batch/sieve values have numerator < denominator.
        # Excel date coercion can flip day/month for slash tokens; canonicalize.
        match = _SLASH_PAIR_RE.match(value)
        if not match:
            return value
        left = int(match.group(1))
        right = int(match.group(2))
        low, high = sorted((left, right))
        return f"{low}/{high}"

    @staticmethod
    def _to_decimal(value: object, default: Decimal) -> Decimal:
        if value is None or value == "":
            return default
        return Decimal(str(value))

    @staticmethod
    def _optional_decimal(value: object) -> Decimal | None:
        if value is None or value == "":
            return None
        return Decimal(str(value))

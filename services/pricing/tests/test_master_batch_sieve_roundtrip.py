from __future__ import annotations

import os
from pathlib import Path

from openpyxl import load_workbook

from app.pricing_engine import DEFAULT_WORKBOOK_PATH, KarigarPricingEngine

TARGET_VALUES = [
    "1/4",
    "1/5",
    "1/3",
    "3/8",
    "1/2",
    "5/8",
    "3/4",
    "9/10",
    "2-6.5",
    "6.5-11",
    "11-14",
    "14-16",
    "-2",
]


def _workbook_path() -> Path:
    return Path(os.getenv("KARIGAR_PRICING_WORKBOOK_PATH", DEFAULT_WORKBOOK_PATH))


def test_master_batch_sieve_values_roundtrip_through_normalizers() -> None:
    """
    Loads the real workbook and prints raw + normalized values so we can inspect
    exactly what openpyxl returns before and after normalization.
    """
    workbook_path = _workbook_path()
    assert workbook_path.exists(), (
        f"Workbook not found at '{workbook_path}'. Set KARIGAR_PRICING_WORKBOOK_PATH."
    )

    wb_raw = load_workbook(filename=str(workbook_path), data_only=False)
    ws_master_raw = wb_raw["Master"]

    try:
        raw_cells: list[object] = []
        for row in range(2, ws_master_raw.max_row + 1):
            raw_cells.append(ws_master_raw.cell(row=row, column=2).value)  # BATCH
            raw_cells.append(ws_master_raw.cell(row=row, column=6).value)  # Sieve Sizes

        assert raw_cells, "Expected Master sheet cells from columns B and F."

        engine = KarigarPricingEngine(workbook_path=str(workbook_path))
        for expected in TARGET_VALUES:
            expected_normalized = engine._normalize_batch_or_sieve(  # noqa: SLF001
                engine._stringify(expected)  # noqa: SLF001
            )

            matched_raw = None
            matched_normalized = None
            for raw in raw_cells:
                normalized = engine._normalize_batch_or_sieve(  # noqa: SLF001
                    engine._stringify(raw)  # noqa: SLF001
                )
                if normalized == expected_normalized:
                    matched_raw = raw
                    matched_normalized = normalized
                    break

            print(
                f"target={expected!r} "
                f"raw={matched_raw!r} "
                f"raw_type={type(matched_raw).__name__ if matched_raw is not None else 'None'} "
                f"normalized={matched_normalized!r}"
            )
            assert (
                matched_raw is not None
            ), f"Could not find workbook B/F cell round-tripping to {expected!r}"
            assert matched_normalized == expected_normalized
    finally:
        wb_raw.close()
